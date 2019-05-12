import os
import conf
import xlrd
from openpyxl import load_workbook
import traceback


class ExcelHelper(object):
    def __init__(self, **kwargs):
        # 初始化参数
        self.source_table = kwargs.get("source_table")
        self.source_sheet = kwargs.get("source_sheet")
        self.title_rowx = kwargs.get("title_rowx")
        self.filter_date = kwargs.get("filter_date")
        self.filter_department = kwargs.get("filter_department")
        self.target_table_path = kwargs.get("target_table_path")
        self.target_sheet = kwargs.get("target_sheet")

    def read_data(self):
        try:
            # 1、读取数据
            # 打开表
            rd = xlrd.open_workbook(self.source_table)
            # 打开sheet
            rtb = rd.sheet_by_name(self.source_sheet)
            # 获取sheet的行数
            tb_rows = rtb.nrows
            # 现在标题
            # 序号	日期	大类	物资名称	规格型号	部件号	单位	数量	单价（SR)	金额	（单价）人民币	计划单位	备注	合同号	计划明细号	物资编码	备注	验收单号	验收金额
            # 现在标题
            # 序号	大类	名称	规格型号	单位	数量	单价RMB	备注	计划明细号	备注	日期
            # 将数字转换为日期
            # 数据汇总
            department_data = {}
            for row in range(self.title_rowx, tb_rows):
                row_list = []
                date = xlrd.xldate_as_datetime(rtb.cell(row, 1).value, 0).strftime("%Y/%m/%d")
                # begin
                # big_type
                row_list.append(int(rtb.row_values(row)[2]))
                # name
                row_list.append(rtb.row_values(row)[3])
                # version
                row_list.append(rtb.row_values(row)[4])
                # unit
                row_list.append(rtb.row_values(row)[6])
                # num
                row_list.append(int(rtb.row_values(row)[7]))
                # price_rmb
                row_list.append(round(float(rtb.row_values(row)[10]), 2))
                # note1
                row_list.append(rtb.row_values(row)[12] if rtb.row_values(row)[12] else "note1")
                # plan_num
                row_list.append(rtb.row_values(row)[14])
                # note2
                row_list.append(rtb.row_values(row)[16] if rtb.row_values(row)[16] else "")
                # date
                row_list.append(date)
                # end

                # 将所有数据放到department_data这个字典中
                department = rtb.row_values(row)[11]
                if department != "":
                    department_date = department + "@" + date
                    if department_date not in department_data:
                        department_data[department_date] = []
                    department_data[department_date].append(row_list)

            # 使用部门和日期进行过滤数据
            filter_data = {}
            for k, v in department_data.items():
                for depart in self.filter_department:
                    if k == str(depart) + "@" + str(self.filter_date):
                        filter_data[k.split("@")[0]] = v

            return filter_data
        except Exception as e:
            print(traceback.format_exc())

    def write_data(self, data):
        try:
            """
           1、如果没有表格，就创建表格和表头 
           2、如果有表格，就追加数据
           """
            # 创建目标文件路径，如果不存在，就新建
            if not os.path.exists(os.path.dirname(self.target_table_path)):
                os.makedirs(self.target_table_path)
            # 查看文件是否存在, 如果不存在就打印不存在信息
            file_name = self.target_table_path
            print("data is %s" % data)
            if os.path.exists(file_name):
                # 文件存在
                # 判断sheet是否存在
                rt = xlrd.open_workbook(file_name)
                sheets = rt.sheet_names()
                # 使用openpyxl打开目标工作簿，目的是保存为xlsx格式的
                new_rt = load_workbook(file_name)
                filter_department_sheet = ["sino-" + x for x in data.keys()]
                for depart in filter_department_sheet:
                    if depart in sheets:
                        # 部门id： 19
                        depart_id = depart.split("-")[1]
                        # 获取目标sheet的句柄
                        rtb = rt.sheet_by_name(depart)
                        # 获取原表格行数
                        rtb_nrows = rtb.nrows
                        # 获取原表格最后一行的id
                        nid = rtb.row_values(rtb_nrows - 1)[0]
                        # 使用openpyxl打开sheet
                        wtb = new_rt[depart]

                        old_data = []
                        # 获取老数据
                        for i in range(self.title_rowx + 2, rtb_nrows):
                            old_data.append(rtb.row_values(i)[1:])
                        print("depart is %s" % depart)
                        print("old_data is %s, len is %s" % (old_data, len(old_data)) )
                        print(data[depart_id])

                        # 循环旧数据，删除新数据中重复的数据
                        for item in old_data:
                            if item in data[depart_id]:
                                data[depart_id].remove(item)

                        # 给每个sheet中插入数据
                        for item in data[depart_id]:
                            # 顺序生成序号
                            if nid != "序号":
                                item.insert(0, nid+1+data[depart_id].index(item))
                            else:
                                item.insert(0, 1+data[depart_id].index(item))
                            # 插入数据
                            wtb.append(item)
                    else:
                        print("no such sheet: %s,please add." % depart)
                # 保存文件
                new_rt.save(file_name)
            else:
                print("文件不存在: %s" % file_name)
        except Exception as e:
            print(traceback.format_exc())


if __name__ == "__main__":
    """
    1、python中都是从0开始计数的
    2、source_table:源数据表
    3、source_sheet:源数据表中的sheet名字
    4、title_rowx:源数据表中标题所在的行数
    5、filter_date:源数据表中使用此参数进行日期过滤
    6、filter_department:源数据表中使用此参数进行部门过滤
    7、target_table_path:目标保存数据表的目录
    8、target_sheet:目标数据保存sheet
    """
    # -----------------------------------------------
    """
    需修改的参数
    """
    source_table = conf.source_table
    source_sheet = conf.source_sheet
    title_rowx = conf.title_rowx
    target_table_path = conf.target_table_path
    departments = conf.departments
    filter_date = conf.filter_date
    # -----------------------------------------------
    try:
        eh = ExcelHelper(source_table=source_table,
                         source_sheet=source_sheet,
                         title_rowx=title_rowx,
                         filter_date=filter_date,
                         filter_department=departments,
                         target_table_path=target_table_path
                         )
        data = eh.read_data()
        if len(data) == 0:
            print("今日[%s]无数据，不需要生成." % filter_date)
        else:
            eh.write_data(data)
            print("---   今日[%s]数据生成完成." % filter_date)
    except Exception as e:
        # print(str(e))
        print(traceback.format_exc())
